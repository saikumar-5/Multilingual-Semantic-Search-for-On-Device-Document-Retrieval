"""
Excel Parser using openpyxl.
Extracts text from all cells across all sheets.

CO1 Alignment: Document Representation - extracting text from spreadsheets.
"""

from pathlib import Path
import logging

logger = logging.getLogger(__name__)


class ExcelParser:
    """Extract text from .xlsx files."""

    def extract(self, file_path: Path) -> dict:
        """
        Extract text from an Excel file.
        Reads all cells from all sheets and concatenates them.

        Returns:
            dict with keys: 'text', 'metadata', 'file_path', 'file_type'
        """
        from openpyxl import load_workbook

        file_path = Path(file_path)
        result = {
            "file_path": str(file_path),
            "file_name": file_path.name,
            "file_type": "excel",
            "text": "",
            "metadata": {},
        }

        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)
            all_text = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_rows = []
                for row in ws.iter_rows(values_only=True):
                    cell_values = [
                        str(cell).strip()
                        for cell in row
                        if cell is not None and str(cell).strip()
                    ]
                    if cell_values:
                        sheet_rows.append(" | ".join(cell_values))

                if sheet_rows:
                    all_text.append(f"[Sheet: {sheet_name}]")
                    all_text.extend(sheet_rows)

            wb.close()
            result["text"] = "\n".join(all_text)
            result["metadata"] = {"sheet_count": len(wb.sheetnames)}

        except Exception as e:
            logger.error(f"Error parsing Excel {file_path}: {e}")
            result["text"] = ""

        return result
